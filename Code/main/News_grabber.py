from telethon.sync import TelegramClient, connection, events
from telethon.sessions import StringSession
import asyncio
from datetime import date, datetime, timedelta
from time import sleep
import pymorphy2
import psycopg2
from sqlalchemy import create_engine
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from xlsxwriter import Workbook
import os, os.path
import python_socks

string = os.getenv('string')
ip = os.getenv('ip')
port = os.getenv('port')

client = TelegramClient(StringSession(string), api_id=os.getenv('API_ID'), api_hash= os.getenv('API_HASH'),proxy=(python_socks.ProxyType.HTTP, ip, port))

donor = '''['ria_realty', 'fontankaspb', 'nedvizha', 'realestate_rf', 'Jelezobetonniyzames', 'banksta', 'ruarbitr',
         'posadky', 'moscowtop', 'domostroy_channel', 'vchkogpu', 'propertyinsider', 'filatofff', 'cotlovan_contrust',
         'belaya_kaska', 'Mos_stroi', 'domtech', 'lietomerealty', 'atsogoev', 'novostroyman', 'me',
         'trubapodneglinnoy', 'riskovik', 'wearestroyka', 'soyzsmet', 'riskovik2', 'real_estate', 'everydayproperty',
         'kompr', 'zmysl', 'chpmoscow', 'zarbitrazhy', 'kleveta']'''

own = os.getenv('own')

m = pymorphy2.MorphAnalyzer()
# функция нормализации текста для каждого слова к первое склонение ед.число
def normalize(s1):
    sent = []
    stroka = s1.replace('?', '')
    stroka = stroka.replace('!', '')
    stroka = stroka.replace('.', '')
    stroka = stroka.replace(',', '')

    s = stroka.lower().split()

    for i in s:
        p = m.parse(i)[0]
        sent.append(p.normal_form)
    s2 = ' '.join(sent)

    return s2


file = os.getenv('EXCEL_FILE_PATH')

def connection():
    dbpass = os.getenv('DBPASS')
    dbname = os.getenv('DBNAME')
    dbuser = os.getenv('DBUSER')
    dbhost = os.getenv('DBHOST')
    dbport = os.getenv('DBPORT')

    con = psycopg2.connect(database=dbname, user = dbuser, password = dbpass,
                           host = dbhost, port = dbport)

    cursor = con.cursor()

    engine = create_engine(f"postgresql+psycopg2://{dbuser}:{dbpass}@{dbhost}:{dbport}/{dbname}?charset=utf8mb4", echo = False)

    return con, cursor, engine

# Фильтры
def fillta(file):

    x = []
    y = []

    wb = openpyxl.load_workbook(file, data_only=True)
    wb.active = 0
    ws = wb.active
    first_filters_list = ws['A2':'A'+str(ws.max_row)]
    second_filters_list = ws['B2':'B'+str(ws.max_row)]

    for row in first_filters_list:
        for cell in row:
            if cell.value is None:
                break
            else:
                x.append(str(cell.value).lower())

    for row in second_filters_list:
        for cell in row:
            if cell.value is None:
                break
            else:
                y.append(str(cell.value).lower())

    # функция нормализации текста для каждого слова к первое склонение ед.число
    def normalize1(s1):
        sent = []
        stroka = s1.replace('?', '')
        stroka = stroka.replace('!', '')
        stroka = stroka.replace('.','')
        stroka = stroka.replace(',','')

        s = stroka.lower().split()

        for i in s:
            p = m.parse(i)[0]
            sent.append(p.normal_form)
        s2 = ' '.join(sent)

        return s2

    lenx = len(x)
    leny = len(y)

    # Создаем список из фильтров, в списке у заменяем пробелы на -, чтобы слить

    total_list = []
    for i in range(0, lenx):
        for j in range(0, leny):
            total_list.append(x[i] +' '+ y[j].replace(' ', '-'))


    lent = len(total_list)

    # По пробелам разделяем слова на отдельные элементы списка
    global_list = []
    for i in range(0,lent):
        global_list.append(total_list[i].split())

    # Обратно заменяем - на пробелы
    for i in range(len(global_list)):
        for j in range(len(global_list[i])):
            global_list[i][j] = normalize1(global_list[i][j].replace('-', ' '))

    # Выбираем позитивные и негативные слова
    wb.active = 1
    ws = wb.active
    cola = ws['A2':'A' + str(ws.max_row)]
    colb = ws['B2':'B' + str(ws.max_row)]
    cole = ws['E2':'E' + str(ws.max_row)]
    cold = ws['D2':'D' + str(ws.max_row)]
    negatives = []
    for row in cola:
        for cell in row:
            negatives.append(cell.value)

    positives = []
    for row in colb:
        for cell in row:
            positives.append(cell.value)
    positive_smiles0 = []
    for row in cold:
        for cell in row:
            positive_smiles0.append(cell.value)

    negative_smiles0 = []
    for row in cole:
        for cell in row:
            negative_smiles0.append(cell.value)

    print(global_list)
    return global_list, negatives, positives, positive_smiles0, negative_smiles0


global_list, negatives, positives, positive_smiles0, negative_smiles0 = fillta(file)


positive_smiles = []
negative_smiles = []
negatives_list = []
positives_list = []

# Убираем None, переводим в нижний регистр
for i in negatives:
    if i == None:
        pass
    else:
        negatives_list.append(normalize(i.lower()))

for i in positives:
    if i == None:
        pass
    else:
        positives_list.append(normalize(i.lower()))

for i in positive_smiles0:
    if i != None:
        positive_smiles.append(i)

for i in negative_smiles0:
    if i != None:
        negative_smiles.append(i)


@client.on(events.MessageEdited(eval(donor)))
@client.on(events.NewMessage(eval(donor)))
async def main(event):
    try:
        x = event.message.to_dict() # переводим ивент в словарь
        # Выделяем id сообщения
        message_id = event.message.id
        # for elem in x:
        #     print("%s -> %s" % (elem, x[elem])) ------ На случай, если нужно посмотреть в удобном формате словарь сообщения
        # Выделяем username
        username = await event.get_sender()

        username = username.username

        # Формируем ссылку на пост

        link = f'https://t.me/{username}/{message_id}'

        # Выделяем имя группы (как в названии)
        channel_name = await event.get_sender()
        try:
            channel_name = channel_name.title
        except Exception as e :
            channel_name = None

        # Выделяем, от кого было переслано сообщение
        try:
            replied_from = event.fwd_from.from_name
        except Exception as e :
            replied_from = None

        # Выделяем текст сообщения
        if event.message.message == '':
            finaltext = 'No caption'
        else:
            finaltext = event.message.message


        # Количество просмотров + пересылок поста

        get_messages = await client.get_messages(username, 25)
        tw_4 = get_messages[24]
        views = tw_4.views
        forwards = tw_4.forwards
        mesid50 =tw_4.id



        # Время публикации

        posted_timestamp = event.date + timedelta(hours=3)
        posted_timestamp = posted_timestamp.strftime('%Y-%m-%d %H:%M:%S')

        # Хештеги
        text = normalize(finaltext)
        txt = text.lower()
        txt0 = txt.split()
        txt1 = finaltext.split()

        str_hashtags = None

        if '#' in text:
            hashtagsdb = []
            str_hashtags = ''

            for i in txt1:
                if '#' in i:
                    hashtagsdb.append(i)
            for i in hashtagsdb:
                if i != hashtagsdb[-1]:
                    str_hashtags += str(i) + ', '
                else:
                    str_hashtags += str(i)

        # Проверка на рекламу

        if ('clc.to' in str(event.message) or '#реклама' in str(event.message) or 'InlineKeyboardMarkup' in str(event.message)):
            ad = 1
        elif 'MessageEntityMention' in str(x):
            if str(username).lower() not in str(finaltext).lower():
                ad = 1
            else:
                ad = 0

        elif 'MessageMediaWebPage' in str(x):
            if str(username).lower() not in str(event.message.media.webpage.url).lower():
                ad = 1
            else:
                ad = 0

        else:
            ad = 0

        # Считаем позитивные и негативные слова
        bad = 0
        good = 0
        good_smiles = 0
        bad_smiles = 0


        con, cursor, engine = connection() #соединение с БД

        # Создаем список наших клиентов
        sql = '''select * from parsing.monitoring_clients'''

        df = pd.read_sql(sql,con)


        final_lst = df['sub'].to_list()
        final0 = [i for i in final_lst if i is not None and i != 'None']
        final = [i.strip() for i in final0]
        final = set(final)


        strx2 = finaltext.split() # разделяем сообщение в список по словам
        final_lower = finaltext.lower() # В другой версии в нижний регистр
        # Список необходимых слов одним словом ранее, чем наш агент
        string = '''общество предприятие фирма организация компания братия команда шатия холдинг товарищество шайка-лейка набор мафия система хавира шарага шарашка шаражка штабель'''
        string2 = string.lower().split()

        
        counter = 0
        match = None
        
        for i in final:
            if ' ' in i and len(i)>=4:
                if i in final_lower:
                    counter +=1
                    match = i

            else:
                for j in strx2:
                    if j==i:
                        index = strx2.index(j)
                        for sos in string2:
                            if sos == normalize(strx2[index-1]):
                                counter +=1
                                match = j
                                

        now = datetime.now() + timedelta(hours=3) # текущее время
        now = now.strftime("%Y-%m-%d %H:%M:%S")

        # Устанавливаем счетчик количества просмотров поста 50 сообщений назад
        cursor.execute("UPDATE parsing.telegram_news_parser SET lviews = %s, forwards = %s where message_id = %s and username = %s",
                    (views, forwards, mesid50, username))
        con.commit()

        # Проверяем запись в БД
        cursor.execute("select message_id from parsing.telegram_news_parser where message_id = %s and username = %s", (message_id, username))
        if not cursor.fetchall():
            # Cчитаем негативн и позитив
            # Smiles
            for pos_smile in positive_smiles:
                for letter in finaltext:
                    if letter == pos_smile:
                        good_smiles += 1

            for neg_smile in negative_smiles:
                for letter in finaltext:
                    if letter == neg_smile:
                        bad_smiles += 1

            for nega in negatives_list:
                for word in txt0:
                    if nega == word:
                        bad += 1

            for pos in positives_list:
                for word in txt0:
                    if pos == word:
                        good += 1

            if good_smiles > 0 or bad_smiles > 0:
                if good_smiles > 0 and bad_smiles == 0:
                    emotional = 'Positive'
                elif bad_smiles > 0 and good_smiles == 0:
                    emotional = 'Negative'
                elif good_smiles / (good_smiles + bad_smiles) < 0.3:
                    emotional = 'Negative'
                elif bad_smiles / (good_smiles + bad_smiles) < 0.3:
                    emotional = 'Positive'
                else:
                    if bad == 0 and good == 0:
                        emotional = 'Neutral'
                    elif bad > 0 and good == 0:
                        emotional = 'Negative'
                    elif good > 0 and bad == 0:
                        emotional = 'Positive'
                    elif good / (good + bad) < 0.5:
                        emotional = 'Negative'
                    elif bad / (good + bad) < 0.3:
                        emotional = 'Positive'
                    else:
                        emotional = "Undefined"
            else:
                if bad == 0 and good == 0:
                    emotional = 'Neutral'
                elif bad > 0 and good == 0:
                    emotional = 'Negative'
                elif good > 0 and bad == 0:
                    emotional = 'Positive'
                elif good / (good + bad) < 0.5:
                    emotional = 'Negative'
                elif bad / (good + bad) < 0.3:
                    emotional = 'Positive'
                else:
                    emotional = "Undefined"

            cursor.execute(
                '''insert into parsing.telegram_news_parser (message_id, username, channel_name, replied_from, replied_from_name,
                message_text, posted_timestamp, activity_date, emotional, is_ad, hashtags, link, finded_agent, clients_cnt)
                values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                (message_id, username, channel_name, replied_from, None, finaltext, posted_timestamp,
                now, emotional, ad, str_hashtags, link, match, counter))
            con.commit()

            for key in global_list:
                if key[0] in txt and key[1] in txt:
                    await client.forward_messages(own, event.message)
                    await client.send_message(own, 'Предыдущее сообщение найдено по фильтру ' + str(key))
                    return

        else: #Если запись уже есть в БД, то обновляем ее, обновляя сообщение, добавляя время изменения, изменяя счетчик изменений
            cursor.execute(
                "select message_text from parsing.telegram_news_parser where message_id = %s and username = %s order by activity_date desc limit 1",
                (message_id, username))

            comparison = cursor.fetchall()
            comparison_text = comparison[0][0]

            if comparison_text != finaltext:

                len_diff = len(finaltext) - len(comparison_text)

                edited_timestamp = now

                for pos_smile in positive_smiles:
                    for letter in finaltext:
                        if letter == pos_smile:
                            good_smiles += 1

                for neg_smile in negative_smiles:
                    for letter in finaltext:
                        if letter == neg_smile:
                            bad_smiles += 1

                for nega in negatives_list:
                    for word in txt0:
                        if nega == word:
                            bad += 1

                for pos in positives_list:
                    for word in txt0:
                        if pos == word:
                            good += 1

                if good_smiles > 0 or bad_smiles > 0:
                    if good_smiles > 0 and bad_smiles == 0:
                        emotional = 'Positive'
                    elif bad_smiles > 0 and good_smiles == 0:
                        emotional = 'Negative'
                    elif good_smiles / (good_smiles + bad_smiles) < 0.3:
                        emotional = 'Negative'
                    elif bad_smiles / (good_smiles + bad_smiles) < 0.3:
                        emotional = 'Positive'
                    else:
                        if bad == 0 and good == 0:
                            emotional = 'Neutral'
                        elif bad > 0 and good == 0:
                            emotional = 'Negative'
                        elif good > 0 and bad == 0:
                            emotional = 'Positive'
                        elif good / (good + bad) < 0.5:
                            emotional = 'Negative'
                        elif bad / (good + bad) < 0.3:
                            emotional = 'Positive'
                        else:
                            emotional = "Undefined"
                else:
                    if bad == 0 and good == 0:
                        emotional = 'Neutral'
                    elif bad > 0 and good == 0:
                        emotional = 'Negative'
                    elif good > 0 and bad == 0:
                        emotional = 'Positive'
                    elif good / (good + bad) < 0.5:
                        emotional = 'Negative'
                    elif bad / (good + bad) < 0.3:
                        emotional = 'Positive'
                    else:
                        emotional = "Undefined"

                cursor.execute("select edit_counter from parsing.telegram_news_parser where message_id = %s and username = %s order by activity_date desc limit 1",
                            (message_id,username))
                edit_counter = cursor.fetchall()
                edit_counter = edit_counter[0][0]

                if edit_counter is None:

                    cursor.execute(
                        '''insert into parsing.telegram_news_parser (message_id, username, channel_name, replied_from, replied_from_name,
                    message_text, posted_timestamp, edited_timestamp, activity_date, emotional, is_ad, hashtags, edit_counter, len_diff, link, finded_agent, clients_cnt) 
                    values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                        (message_id, username, channel_name, replied_from, None, finaltext, posted_timestamp,
                        edited_timestamp, now, emotional, ad,
                        str_hashtags, 1, len_diff, link, match, counter))

                else:

                    cursor.execute(
                        '''insert into parsing.telegram_news_parser (message_id, username, channel_name, replied_from, replied_from_name,
                    message_text, posted_timestamp, edited_timestamp, activity_date, emotional, is_ad, hashtags, edit_counter, len_diff, link, finded_agent, clients_cnt) 
                    values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s, %s)''',
                        (message_id, username, channel_name, replied_from, None, finaltext, posted_timestamp,
                        edited_timestamp, now, emotional, ad,
                        str_hashtags, edit_counter + 1, len_diff, link, match, counter))

                con.commit()

    except FloodWaitError as e:
        print('Flood waited for', e.seconds)
        time.sleep(3)

client.start()
client.run_until_disconnected()
